import openpyxl
from openpyxl.styles import NamedStyle

# Crear estilo para los porcentajes
percentage_style = NamedStyle(name="percentage_style")

# Datos completos que serán ingresados en el archivo Excel (basados en los que me pasaste)
data = [
    ['VRPTW1', 191.8136197786562, 98.99, 93.77, 3, 3, 0.00, 8802],
    ['VRPTW2', 215.54256527689375, 141.82, 51.98, 2, 1, 100.00, 15640],
    ['VRPTW3', 618.3299155462686, 276.78, 123.40, 8, 2, 300.00, 7770],
    ['VRPTW4', 524.593091000854, 276.78, 89.53, 2, 1, 100.00, 17942],
    ['VRPTW5', 462.15594675837383, 178.41, 159.05, 4, 3, 33.33, 6710],
    ['VRPTW6', 467.95268098026867, 178.41, 162.30, 2, 1, 100.00, 15044],
    ['VRPTW7', 363.2468004115909, 195.66, 85.65, 5, 5, 0.00, 16442],
    ['VRPTW8', 444.96097420146134, 264.11, 68.48, 2, 2, 0.00, 49231],
    ['VRPTW9', 1079.7343441633047, 417.38, 158.69, 12, 4, 200.00, 11402],
    ['VRPTW10', 1021.4121183836749, 417.38, 144.72, 2, 1, 100.00, 40026],
    ['VRPTW11', 962.0414969341415, 315.74, 204.70, 9, 5, 80.00, 16943],
    ['VRPTW12', 948.5563805751098, 315.74, 200.43, 3, 1, 200.00, 35296],
    ['VRPTW13', 828.9368669428338, 417.30, 98.64, 10, 10, 0.00, 68795],
    ['VRPTW14', 591.5565566715014, 492.47, 20.12, 3, 3, 0.00, 197839],
    ['VRPTW15', 1674.571622374453, 562.26, 197.83, 20, 8, 150.00, 57912],
    ['VRPTW16', 1261.7111072766616, 562.26, 124.40, 12, 2, 500.00, 78706],
    ['VRPTW17', 1767.6902824607505, 564.00, 213.42, 17, 9, 88.89, 59721],
    ['VRPTW18', 1376.2400501275126, 564.00, 144.02, 9, 2, 350.00, 89903]
]








# Crear un archivo Excel
wb = openpyxl.Workbook()
ws = wb.active

# Definir las columnas
columns = ["Filename", "Total Distance", "Lower Bound Distance", "GAP Distance (%)", 
           "Actual Routes", "Lower Bound Routes", "GAP Routes (%)", "Execution Time (ms)"]
ws.append(columns)

# Agregar los datos fila por fila
for row in data:
    ws.append(row)

# Aplicar formato de porcentaje a las columnas correspondientes
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):  # Columna de GAP Distance (%)
    for cell in row:
        cell.value = cell.value / 100  # Convertir el valor a decimal

for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):  # Columna de GAP Routes (%)
    for cell in row:
        cell.value = cell.value / 100  # Convertir el valor a decimal

# Guardar el archivo Excel
output_path = "prueba3800.xlsx"
wb.save(output_path)

