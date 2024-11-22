## Juan Fernando Riascos Goyes
## VRPTW Solution Improvement using Strict VNS with Multiple Neighborhoods

import os
import math
import numpy as np
import time
from openpyxl import load_workbook, Workbook
from multiprocessing import Pool, cpu_count
from Lecture import Nodo, save_to_excel, plot_routes, read_txt_file
from Feasibility_and_LB import lower_bound_mst, lower_bound_routes, is_feasible
import random
from numba import njit, prange

random.seed(4)
np.random.seed(4)

def dist(node1, node2):
    return math.sqrt((node1.x_cord - node2.x_cord) ** 2 + (node1.y_cord - node2.y_cord) ** 2)

def calculate_travel_times(nodes):
    n = len(nodes)
    times = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            times[i][j] = dist(nodes[i], nodes[j])
    return times

def calculate_route_distance(route, times):
    distance = 0.0
    for i in range(len(route) - 1):
        distance += times[route[i].index][route[i + 1].index]
    return distance

def calculate_total_distance(routes, times):
    return sum(calculate_route_distance(route, times) for route in routes)

def is_route_feasible(route, capacity, times):
    feasible_route = [route[0]]
    for node in route[1:]:
        if is_feasible(feasible_route, node, capacity, times):
            feasible_route.append(node)
        else:
            return False
    return True



def swap_between_routes_best(routes, times, capacity):
    best_routes = [route.copy() for route in routes]
    best_distance = calculate_total_distance(best_routes, times)
    improved = False

    for i in range(len(routes)):
        for j in range(i + 1, len(routes)):
            route1 = routes[i]
            route2 = routes[j]

            customers1 = route1[1:-1]
            customers2 = route2[1:-1]

            for idx1, cust1 in enumerate(customers1):
                for idx2, cust2 in enumerate(customers2):
                    temp_route1 = route1.copy()
                    temp_route2 = route2.copy()

                    temp_route1[idx1 + 1] = cust2
                    temp_route2[idx2 + 1] = cust1

                    if (is_route_feasible(temp_route1, capacity, times) and
                        is_route_feasible(temp_route2, capacity, times)):
                        temp_routes = routes.copy()
                        temp_routes[i] = temp_route1
                        temp_routes[j] = temp_route2
                        temp_distance = calculate_total_distance(temp_routes, times)
                        if temp_distance + 1e-6 < best_distance:
                            best_routes = temp_routes
                            best_distance = temp_distance
                            improved = True
    return best_routes, best_distance, improved

def relocate_between_routes_best(routes, times, capacity):
    """
    Función mejorada para reubicar clientes entre rutas de manera más agresiva.
    Intenta mover secuencias de clientes de una ruta a otra.
    """
    best_routes = [route.copy() for route in routes]
    best_distance = calculate_total_distance(best_routes, times)
    improved = False

    for i in range(len(routes)):
        for j in range(len(routes)):
            if i == j:
                continue

            route_from = routes[i]
            route_to = routes[j]

            customers_from = route_from[1:-1]

            # Intentar mover secuencias de diferentes tamaños
            for seq_length in range(1, len(customers_from) + 1):
                for idx_cust in range(len(customers_from) - seq_length + 1):
                    segment = customers_from[idx_cust:idx_cust + seq_length]

                    temp_route_from = route_from[:idx_cust + 1] + route_from[idx_cust + seq_length + 1:]
                    if not is_route_feasible(temp_route_from, capacity, times):
                        continue

                    for k in range(1, len(route_to)):
                        temp_route_to = route_to[:k] + segment + route_to[k:]
                        if is_route_feasible(temp_route_to, capacity, times):
                            temp_routes = routes.copy()
                            temp_routes[i] = temp_route_from
                            temp_routes[j] = temp_route_to
                            temp_distance = calculate_total_distance(temp_routes, times)
                            if temp_distance + 1e-6 < best_distance:
                                best_routes = temp_routes
                                best_distance = temp_distance
                                improved = True
    return best_routes, best_distance, improved


def two_opt_within_route_single(route, times, capacity):
    best_route = route.copy()
    best_distance = calculate_route_distance(best_route, times)
    n = len(route)
    improved = False

    for i in range(1, n - 2):
        for j in range(i + 1, n - 1):
            new_route = route[:i] + route[i:j+1][::-1] + route[j+1:]
            if is_route_feasible(new_route, capacity, times):
                new_distance = calculate_route_distance(new_route, times)
                if new_distance + 1e-6 < best_distance:
                    best_route = new_route
                    best_distance = new_distance
                    improved = True
    return best_route, best_distance, improved

def two_opt_across_routes(routes, times, capacity):
    best_routes = [route.copy() for route in routes]
    best_distance = calculate_total_distance(best_routes, times)
    improved = False

    for i in range(len(routes)):
        for j in range(i + 1, len(routes)):
            route1 = routes[i]
            route2 = routes[j]

            for idx1 in range(1, len(route1) - 1):
                for idx2 in range(1, len(route2) - 1):
                    new_route1 = route1[:idx1] + route2[idx2:]
                    new_route2 = route2[:idx2] + route1[idx1:]

                    if (is_route_feasible(new_route1, capacity, times) and
                        is_route_feasible(new_route2, capacity, times)):
                        temp_routes = routes.copy()
                        temp_routes[i] = new_route1
                        temp_routes[j] = new_route2
                        temp_distance = calculate_total_distance(temp_routes, times)
                        if temp_distance + 1e-6 < best_distance:
                            best_routes = temp_routes
                            best_distance = temp_distance
                            improved = True
    return best_routes, best_distance, improved



def or_opt_within_route_single(route, times, capacity):
    best_route = route.copy()
    best_distance = calculate_route_distance(best_route, times)
    n = len(route)
    improved = False

    for segment_size in range(1, 4):
        for i in range(1, n - segment_size - 1):
            segment = route[i:i+segment_size]
            rest_route = route[:i] + route[i+segment_size:]

            for j in range(1, len(rest_route)):
                new_route = rest_route[:j] + segment + rest_route[j:]
                if is_route_feasible(new_route, capacity, times):
                    new_distance = calculate_route_distance(new_route, times)
                    if new_distance + 1e-6 < best_distance:
                        best_route = new_route
                        best_distance = new_distance
                        improved = True
                        return best_route, best_distance, improved  # Salir inmediatamente
    return best_route, best_distance, improved





def vnd_algorithm(routes, times, capacity, time_limit, start_time):
    """
    Algoritmo VND con criterio de parada basado en tiempo.
    """
    best_routes = [route.copy() for route in routes]
    best_distance = calculate_total_distance(best_routes, times)
    neighborhoods = [
        two_opt_within_route_single,
        or_opt_within_route_single,
        swap_between_routes_best,
        relocate_between_routes_best,
        two_opt_across_routes,
        merge_routes
    ]

    neighborhood_index = 0

    while neighborhood_index < len(neighborhoods):
        current_time = time.time()
        elapsed_time = current_time - start_time
        if elapsed_time >= time_limit:
            # Tiempo límite alcanzado, terminar
            break

        neighborhood = neighborhoods[neighborhood_index]
        improved = False

        if neighborhood in [two_opt_within_route_single, or_opt_within_route_single]:
            # Aplicar movimientos dentro de rutas individuales
            for idx, route in enumerate(best_routes):
                new_route, new_distance, route_improved = neighborhood(route, times, capacity)
                if route_improved and new_distance + 1e-6 < calculate_route_distance(best_routes[idx], times):
                    best_routes[idx] = new_route
                    best_distance = calculate_total_distance(best_routes, times)
                    improved = True

                    # Verificar el tiempo después de cada mejora
                    current_time = time.time()
                    elapsed_time = current_time - start_time
                    if elapsed_time >= time_limit:
                        break
            if elapsed_time >= time_limit:
                break
        else:
            if neighborhood == merge_routes:
                new_routes = neighborhood(best_routes, times, capacity)
                new_distance = calculate_total_distance(new_routes, times)
                neighborhood_improved = (len(new_routes) < len(best_routes) or new_distance + 1e-6 < best_distance)
            else:
                new_routes, new_distance, neighborhood_improved = neighborhood(best_routes, times, capacity)

            if neighborhood_improved and (new_distance + 1e-6 < best_distance or len(new_routes) < len(best_routes)):
                best_routes = [route.copy() for route in new_routes]
                best_distance = new_distance
                improved = True

                # Verificar el tiempo después de cada mejora
                current_time = time.time()
                elapsed_time = current_time - start_time
                if elapsed_time >= time_limit:
                    break

        if improved:
            # Continuar con el mismo vecindario
            neighborhood_index = 0
        else:
            # Pasar al siguiente vecindario
            neighborhood_index += 1

    return best_routes










def merge_routes(routes, times, capacity):
    """
    Función mejorada para fusionar rutas de manera más agresiva.
    Intenta fusionar rutas considerando todos los pares posibles y las inversiones de rutas.
    """
    improved = True
    while improved:
        improved = False
        num_routes = len(routes)
        best_distance = calculate_total_distance(routes, times)
        best_routes = routes.copy()

        for i in range(num_routes):
            for j in range(num_routes):
                if i >= j:
                    continue

                route1 = routes[i]
                route2 = routes[j]

                # Intentar fusionar route1 y route2 directamente
                merged_route = route1[:-1] + route2[1:]
                if is_route_feasible(merged_route, capacity, times):
                    temp_routes = [routes[k] for k in range(num_routes) if k != i and k != j]
                    temp_routes.append(merged_route)
                    temp_distance = calculate_total_distance(temp_routes, times)
                    if temp_distance + 1e-6 < best_distance or len(temp_routes) < len(best_routes):
                        best_routes = temp_routes
                        best_distance = temp_distance
                        improved = True
                        break

                # Intentar fusionar route1 y la inversión de route2
                reversed_route2 = [route2[0]] + route2[1:-1][::-1] + [route2[-1]]
                merged_route = route1[:-1] + reversed_route2[1:]
                if is_route_feasible(merged_route, capacity, times):
                    temp_routes = [routes[k] for k in range(num_routes) if k != i and k != j]
                    temp_routes.append(merged_route)
                    temp_distance = calculate_total_distance(temp_routes, times)
                    if temp_distance + 1e-6 < best_distance or len(temp_routes) < len(best_routes):
                        best_routes = temp_routes
                        best_distance = temp_distance
                        improved = True
                        break

            if improved:
                routes = best_routes
                break  # Reiniciar búsqueda desde el principio

    return routes

def read_solution_from_excel(filename, sheet_name, nodes):
    wb = load_workbook(filename=filename, data_only=True)
    sheet = wb[sheet_name]

    routes = []
    num_vehicles = int(sheet['A1'].value)
    total_distance = float(sheet['B1'].value)
    computation_time = int(sheet['C1'].value)

    row = 2
    for _ in range(num_vehicles):
        num_nodes = int(sheet.cell(row=row, column=1).value)
        route_indices = []

        # Leer ruta
        for col in range(2, 2 + num_nodes + 2):  # +2 para incluir ambos depósitos
            node_index = int(sheet.cell(row=row, column=col).value)
            route_indices.append(node_index)

        # Construir ruta
        route = [nodes[idx] for idx in route_indices]
        routes.append(route)
        row += 1

    wb.close()
    return routes, total_distance, computation_time


def vrptw_improve_solutions(excel_filename, directory_path, output_filename):
    wb_output = Workbook()
    wb_output.remove(wb_output.active)

    execution_times = []
    output_folder = "ACO_VND_Images"

    # Diccionario de tiempos límites por instancia
    time_limits = {
        'VRPTW1': 50,
        'VRPTW2': 50,
        'VRPTW3': 50,
        'VRPTW4': 50,
        'VRPTW5': 50,
        'VRPTW6': 50,
        'VRPTW7': 200,
        'VRPTW8': 200,
        'VRPTW9': 200,
        'VRPTW10': 200,
        'VRPTW11': 200,
        'VRPTW12': 200,
        'VRPTW13': 750,
        'VRPTW14': 750,
        'VRPTW15': 750,
        'VRPTW16': 750,
        'VRPTW17': 750,
        'VRPTW18': 750
    }

    for i in range(1, 19):
        instance_name = f'VRPTW{i}'
        excel_sheet_name = instance_name
        data_filename = f'{directory_path}/{instance_name}.txt'

        # Leer nodos y calcular la matriz de tiempos
        n, Q, nodes = read_txt_file(data_filename)
        times = calculate_travel_times(nodes)

        # Leer solución existente desde Excel
        routes, total_distance, computation_time = read_solution_from_excel(excel_filename, excel_sheet_name, nodes)

        # Obtener el tiempo límite para la instancia actual
        time_limit = time_limits.get(instance_name, 50)  # Por defecto 50 si no está especificado

        # Aplicar VND para mejorar la solución con el tiempo límite
        start_time = time.time()
        improved_routes = vnd_algorithm(routes, times, Q, time_limit, start_time)
        improved_distance = calculate_total_distance(improved_routes, times)
        improvement_time = (time.time() - start_time) * 1000  # Tiempo en milisegundos
        execution_times.append(improvement_time)

        # Calcular cotas inferiores
        depot = nodes[0]
        customers = nodes[1:]
        lb_routes = lower_bound_routes(customers, Q)
        lb_distance = lower_bound_mst(depot, customers, times)

        actual_routes = len(improved_routes)
        gap_routes = max(((actual_routes - lb_routes) / lb_routes) * 100 if lb_routes > 0 else 0, 0)
        gap_distance = max(((improved_distance - lb_distance) / lb_distance) * 100 if lb_distance > 0 else 0, 0)

        # Mostrar detalles de la solución
        print(f"Improved Solution for {instance_name}:")
        print(f"  - Total Distance = {improved_distance}")
        print(f"  - Lower Bound Distance (MST) = {lb_distance:.2f}")
        print(f"  - GAP Distance = {gap_distance:.2f}%")
        print(f"  - Total Routes = {actual_routes}")
        print(f"  - Lower Bound Routes = {lb_routes}")
        print(f"  - GAP Routes = {gap_routes:.2f}%")
        print(f"  - Time = {improvement_time:.0f} ms\n")

        # Guardar resultados en Excel
        sheet_name = instance_name
        save_to_excel(wb_output, sheet_name, improved_routes, improved_distance, improvement_time, times)

        # Generar gráficas si lo deseas
        plot_routes(improved_routes, f"{instance_name}.txt", output_folder)

    wb_output.save(output_filename)

    total_elapsed_time = sum(execution_times)
    print(f"\nTotal improvement time: {total_elapsed_time:.0f} ms")


if __name__ == '__main__':
    excel_filename = "VRPTW_JuanFernando_ACO.xlsx"  # Archivo Excel con las soluciones existentes
    directory_path = "./Examples"  # Carpeta con los archivos de entrada
    output_filename = "VRPTW_JuanFernando__aco_VND.xlsx"
    vrptw_improve_solutions(excel_filename, directory_path, output_filename)

